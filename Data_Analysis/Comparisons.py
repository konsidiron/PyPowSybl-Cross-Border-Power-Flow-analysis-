import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os  

"""
SPECIFIC IGM COMPARISON OF I,P,Q IN X-LINES/LINES OF OPENLF/UNICORN and v, theta for NODES/X-Nodes 
"""

def get_user_inputs():
    """
    Function to get user inputs for the comparison process.
    """
    # Prompt user to enter the necessary details
    timestamps = input("Enter the hours (comma-separated ex. 0030,0130 ... , or leave blank for default 0030-2330): ")
    if not timestamps:
        timestamps = ['0030', '0130', '0230', '0330', '0430', '0530', '0630', '0730', '0830', '0930', '1030', '1130',
                 '1230', '1330', '1430', '1530', '1630', '1730', '1830', '1930', '2030', '2130', '2230', '2330']
    else:
        timestamps = timestamps.split(',')

    destination_folder = input("Enter the base folder path where the load flows are located: ")
    destination_folder_1 = input("Enter the folder path where comparison results will be saved: ")
    Date = input("Enter the date (in YYYYMMDD format): ")
    File_type = input("Enter the file type (e.g., 'FO3'): ")
    country_code = input("Enter the country code (e.g., 'GR'): ")
    #Corresponds to different versions of loadflow reports
    numbers = range(0,15) 
    
    return destination_folder, destination_folder_1, Date, File_type, country_code, numbers , timestamps

def calculate_line_differencies(merged_df):
    """
    Calculate the differences and percentage differences for line data (current, active power, reactive power) and cleans unnecessary data. 
    """
    #Clean data from near to zero values.
    threshold = 1e-2
    merged_df[['I_UNICORN', 'I_OPENLF', 'P_UNICORN', 'P_OPENLF', 'Q_UNICORN', 'Q_OPENLF']] = merged_df[['I_UNICORN', 'I_OPENLF', 'P_UNICORN', 'P_OPENLF', 'Q_UNICORN', 'Q_OPENLF']].applymap(lambda x: 0 if abs(x) < threshold else x)

    merged_df['I_diff'] = merged_df['I_UNICORN'] - merged_df['I_OPENLF']
    merged_df['P_diff'] = merged_df['P_UNICORN'] - merged_df['P_OPENLF']
    merged_df['Q_diff'] = merged_df['Q_UNICORN'] - merged_df['Q_OPENLF']
    
    # Absolute differences
    merged_df['I_diff_abs'] = merged_df['I_diff'].abs()
    merged_df['P_diff_abs'] = merged_df['P_diff'].abs()
    merged_df['Q_diff_abs'] = merged_df['Q_diff'].abs()
    
    # Percentage differences
    merged_df['I_diff_pct'] = (merged_df['I_diff'].abs() / merged_df['I_UNICORN'].abs()) * 100
    merged_df['P_diff_pct'] = (merged_df['P_diff'].abs() / merged_df['P_UNICORN'].abs()) * 100
    merged_df['Q_diff_pct'] = (merged_df['Q_diff'].abs() / merged_df['Q_UNICORN'].abs()) * 100

    # CLEANING DATA FROM DIVISION WITH NEAR TO ZERO OR ZERO VALUES. Their respectively unicorns and openlf's values have absolute differenccies near to zero so instead of inf we will have zero values
    merged_df.replace([np.inf, -np.inf], np.nan, inplace=True) 
    columns = ['I_UNICORN', 'I_OPENLF', 'P_UNICORN', 'P_OPENLF', 'Q_UNICORN', 'Q_OPENLF','I_diff_abs', 'P_diff_abs' ,'Q_diff_abs' ,'I_diff_pct', 'P_diff_pct', 'Q_diff_pct']
    # Fill empty cells with zero values
    merged_df[columns] = merged_df[columns].fillna(0)
    columns_to_check = ['I_UNICORN', 'I_OPENLF', 'P_UNICORN', 'P_OPENLF', 'Q_UNICORN', 'Q_OPENLF']
    # Drop rows where I,P,Q UNICORN'S AND OPENLF'S values are zero
    merged_df = merged_df[~(merged_df[columns_to_check] == 0).all(axis=1)]
    
    return merged_df

def calculate_Nodes_differencies(merged_df):
    """
    Calculate the voltage absolute and percentage differences for nodes data (voltage magnitude and angle) and cleans unnecessary data. 
    """
    # Calculate voltage magnitude (U) and angle (theta) differences
    merged_df['U_diff'] = merged_df['U_UNICORN'] - merged_df['U_OPENLF']
    merged_df['theta_diff'] = merged_df['theta_UNICORN'] - merged_df['theta_OPENLF']
    
    # Absolute differences
    merged_df['U_diff_abs'] = merged_df['U_diff'].abs()
    merged_df['theta_diff_abs'] = merged_df['theta_diff'].abs()
    
    # Percentage differences (absolute percentage difference relative to UNICORN)
    merged_df['U_diff_pct'] = (merged_df['U_diff'].abs() / merged_df['U_UNICORN'].abs()) * 100
    merged_df['theta_diff_pct'] = (merged_df['theta_diff'].abs() / merged_df['theta_UNICORN'].abs()) * 100
    
    # Replace inf or NaN values due to division by zero
    merged_df.replace([np.inf, -np.inf], np.nan, inplace=True)
    merged_df.fillna(0, inplace=True)

    return merged_df


def load_data(filepath, sheet_name):
    """
    Load Excel data from a specified sheet.
    """
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error loading {sheet_name} from {filepath}: {e}")
        return None

#Renames columns and cuts id strings in both company's loadflow reports
def rename_lines_data(df1 , df2):
    df1.rename(columns={'Name (mrid)': 'id'}, inplace=True)
    df1.rename(columns={'Terminal number': 'side'}, inplace=True)
    df2.rename(columns={'side_x': 'side'}, inplace=True)
    df1['id'] = df1['id'].astype(str).str[:19]
    df2['id'] = df2['id'].astype(str).str[:19]  

    return df1, df2

def rename_X_lines_data(df1, df2):
    df2.rename(columns={'BUS' : 'Bus' }, inplace=True)
    df1.columns = df1.columns.str.strip()
    df1.rename(columns={'Name (mrid)': 'id'}, inplace=True)
    df1['id'] = df1['id'].astype(str).str[:19]
    df1['Bus'] = df1['Bus'].astype(str).str[:8]

    return df1,df2 

def rename_X_Nodes_data(df1 , df2):
    def process_id(id_str):
                id_str = str(id_str) 
                if id_str.startswith('X'):
                    return id_str[:8] 
                elif len(id_str) > 9 and id_str[9] == 'X':  
                    return id_str[9:17]  
                return id_str 
            
    df1['Name (mrid)'] = df1['Name (mrid)'].astype(str).str[:8]
    df1.rename(columns={'Name (mrid)': 'id'}, inplace=True)
    df2.rename(columns={'boundary_v_mag': 'U', 'boundary_v_angle': 'theta'}, inplace=True) 
    df2['id'] = df2['id'].apply(process_id)
    
    return df1, df2


def rename_Nodes_data(df1, df2):
    df1['Name (mrid)'] = df1['Name (mrid)'].astype(str).str[:8]
    df1.rename(columns={'Name (mrid)': 'Bus'}, inplace=True)
    df2.rename(columns={ 'BUS' : 'Bus', 'v_mag': 'U', 'v_angle': 'theta'}, inplace=True)

    return df1, df2

def merge_common_data(df1, df2, merge_columns, sort_columns = None):
    """
    Merge two DataFrames on common columns and sort them by columns provided.
    """
    common_ids = set(df1[merge_columns[0]]).intersection(set(df2[merge_columns[0]]))
    df1 = df1[df1[merge_columns[0]].isin(common_ids)]
    df2 = df2[df2[merge_columns[0]].isin(common_ids)]
    
    merged_df = pd.merge(df1, df2, on=merge_columns, suffixes=('_UNICORN', '_OPENLF'))
    if sort_columns:
        merged_df.sort_values(by=sort_columns, ascending=[True] * len(sort_columns), inplace=True)
    return merged_df

# Rename final_df's columns for Lines and X-Lines
def final_columns_rename_lines(merged_df , timestamp):
    final_df = merged_df.rename(columns={
                'I_UNICORN': 'I',
                'I_OPENLF': 'I',
                'P_UNICORN': 'P',
                'P_OPENLF': 'P',
                'Q_UNICORN' : 'Q',
                'Q_OPENLF' : 'Q',
                'I_diff_abs': 'I_diff_abs',
                'P_diff_abs': 'P_diff_abs',
                'Q_diff_abs': 'Q_diff_abs',
                'I_diff_pct': 'I_diff_pct',
                'P_diff_pct': 'P_diff_pct',
                'Q_diff_pct' :'Q_diff_pct',
            }) 
    final_df['Timestamp'] = timestamp   
    return final_df , timestamp

#Rename final_df's columns for Nodes and X-Nodes
def final_columns_rename_buses(merged_df , timestamp):
    final_df = merged_df.rename(columns={
                'U_UNICORN': 'U',
                'U_OPENLF': 'U',
                'theta_UNICORN': 'theta',
                'theta_OPENLF': 'theta',
                'U_diff_abs': 'U_diff_abs',
                'theta_diff_abs': 'theta_diff_abs',
                'U_diff_pct': 'U_diff_pct',
                'theta_diff_pct': 'theta_diff_pct'
            })

    final_df['Timestamp'] = timestamp 
    final_df = final_df.dropna(subset=['U'])
    return final_df , timestamp

def make_adjustements_lines_to_excel(output_path, sheet_name):
    header_titles = [
        'ID', 'SIDE', 'UNICORN', 'UNICORN', 'UNICORN', 'OPENLF', 'OPENLF', 'OPENLF',
        'ABSOLUTE DIFFERENCES', 'ABSOLUTE DIFFERENCES', 'ABSOLUTE DIFFERENCES',
        'PERCENTAGE DIFFERENCES', 'PERCENTAGE DIFFERENCES', 'PERCENTAGE DIFFERENCES', 'TIMESTAMP'
    ]
    
    # Open the workbook and the relevant sheet
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    
    # Insert custom headers
    ws.insert_rows(1)  # Insert an empty row at the top
    for col, value in enumerate(header_titles, start=1):
        ws.cell(row=1, column=col, value=value)  # Write the headers
    
    # Save the workbook with the updated headers
    wb.save(output_path)

def make_adjustements_nodes_to_excel(output_path, sheet_name):
    header_titles = ['ID', 'UNICORN', 'UNICORN', 'OPENLF', 'OPENLF', 'ABSOLUTE DIFFERENCES', 'ABSOLUTE DIFFERENCES', 'PERCENTAGE DIFFERENCES' , 'PERCENTAGE DIFFERENCES'  , 'TIMESTAMP']
    # Open the workbook and the relevant sheet
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
   
    ws.insert_rows(1)
    for col, value in enumerate(header_titles, start=1):
        ws.cell(row=1, column=col, value=value)
    
    wb.save(output_path)


def generate_file_paths(timestamp, number, Date, File_type, country_code, destination_folder):
    # Define the paths for the input files
    df1_path = os.path.join(destination_folder, f'{Date}_{timestamp}_{File_type}_{country_code}_{number}_igmLfReport.xlsx') ####sos USER HAS TO FILL THE RIGHT NAME STRUCTURE OF UNICORN'S LOAD FLOW REPORTS (IGMS)
    # df1_path = os.path.join(destination_folder, f'LfReport_{Date}_{timestamp}_{File_type}_{country_code}{number}.xlsx') # FOR CGMS 
    df2_path = os.path.join(destination_folder, f'{Date}_{timestamp}_{File_type}_{country_code}_0_OPENLF_REPORT.xlsx') ###sos USER HAS TO FILL THE RIGHT NAME STRUCTURE OF OPENLF'S LOAD FLOW REPORTS
    print(f"Generated df1_path: {df1_path}")
    print(f"Generated df2_path: {df2_path}")
    # Check if both files exist
    if os.path.exists(df1_path) and os.path.exists(df2_path):
        
        return df1_path, df2_path
    else:
        # Notify the user about the missing files and offer guidance
        missing_files = []
        if not os.path.exists(df1_path):
            missing_files.append(f"'{df1_path}'")
        if not os.path.exists(df2_path):
            missing_files.append(f"'{df2_path}'")
        
        print(f"Warning: The following expected files were not found:\n{', '.join(missing_files)}")
        print("Please ensure the filenames match the expected pattern or adjust the filenames accordingly.")       
        # Return None to indicate missing files
        return None, None
    
def find_highest_version_number(Date, timestamp, numbers, File_type, country_code , destination_folder):
    highest_number = -1

    for number in numbers:
        report_filename = os.path.join(destination_folder, f'{Date}_{timestamp}_{File_type}_{country_code}_{number}_igmLfReport.xlsx')

        if os.path.exists(report_filename):
            if number > highest_number:
                highest_number = number
               
    return highest_number

def process_files_and_accumulate_data(timestamps, numbers, Date, File_type, country_code, destination_folder, destination_folder_1):
    # Use a single output Excel file for all timestamps
    combined_output_path = os.path.join(destination_folder_1, f'combined_results_OpenLF_Unicorn_{Date}.xlsx')
    # Create dictionaries to store data for each category across all timestamps
    all_sheets_data = {'Lines': [], 'X-lines': [], 'Nodes': [], 'X-Nodes': []}
    
    for timestamp in timestamps:
        number = find_highest_version_number(Date, timestamp, numbers, File_type, country_code, destination_folder)
        # Generate file paths
        df1_path, df2_path = generate_file_paths(timestamp, number, Date, File_type, country_code, destination_folder)
            # Check if the paths exist, continue to the next iteration if they don't
        if df1_path and df2_path and os.path.exists(df1_path) and os.path.exists(df2_path):
             
                #Lines
                df1 = load_data(df1_path , 'Line')
                df2 = load_data(df2_path , 'Line')
                df1, df2 = rename_lines_data(df1, df2)
                merged_df = merge_common_data(df1, df2, merge_columns=['id', 'side'], sort_columns=['id', 'side'])
                merged_df = calculate_line_differencies(merged_df)
                columns_to_drop = ['Terminal number', 'Bus ' ,'BUS', 'v_mag', 'v_angle', 'I_limit' , 'Area' ,  'Island number' , 'U' ,'theta',  'Base Voltage' , 'U' ,'theta', 'Bus',   'Imax' , 'loading' , 'Eq. type' , 'State' , 'r' , 'x' , 'side_x' , 'element_type' , 'side_y' , 'name' , 'type' , 'value' , 'acceptable_duration' , 'I_diff' , 'P_diff' , 'Q_diff']   
                merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns] , inplace= True)
                final_df, timestamp = final_columns_rename_lines(merged_df, timestamp)
                # Drop rows where all columns except 'Timestamp' , 'Bus' , 'id' contain zeros
                columns_to_check = final_df.columns.difference(['Timestamp', 'side', 'id'])
                final_df = final_df.loc[~(final_df[columns_to_check] == 0).all(axis=1)]
                all_sheets_data['Lines'].append(final_df)
                
                
                #X-lines
                df1 = load_data(df1_path, 'Line')
                df2 = load_data(df2_path, 'X-Nodes')
                df1, df2 = rename_X_lines_data(df1, df2)
                merged_df = merge_common_data(df1, df2, merge_columns=['id', 'Bus'] , sort_columns=['id', 'Bus'])
                merged_df = calculate_line_differencies(merged_df)
                columns_to_drop = ['Area' , 'Terminal number' , 'v_angle' , 'I_limit' ,'Island number' , 'U' , 'v_mag' , 'v_angle' 'theta' , 'Base Voltage' , 'U' ,'theta',   'side_x' , 'Imax' , 'Unnamed: 0','loading' , 'Eq. type' , 'State' , 'r' , 'x'  , 'element_type' , 'side_y' , 'name' , 'type' , 'value' , 'acceptable_duration' , 'I_diff' , 'P_diff' , 'Q_diff', 'boundary_v_mag' , 'boundary_v_angle' , 'boundary_p' , 'boundary_q']   
                merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns] , inplace= True)
                final_df, timestamp = final_columns_rename_lines(merged_df, timestamp)
                # Drop rows where all columns except 'Timestamp' , 'Bus' , 'id' contain zeros
                columns_to_check = final_df.columns.difference(['Timestamp', 'Bus', 'id'])
                final_df = final_df.loc[~(final_df[columns_to_check] == 0).all(axis=1)]
                all_sheets_data['X-lines'].append(final_df)
                

                #Nodes
                df1 = load_data(df1_path , 'Bus')
                df2 = load_data(df2_path, 'Bus')
                df1, df2 = rename_Nodes_data(df1, df2)
                merged_df = merge_common_data(df1, df2, merge_columns=['Bus'])
                merged_df = calculate_Nodes_differencies(merged_df)
                columns_to_drop = ['Bus type', 'Reference voltage_UNICORN', 'Pgen_UNICORN' , 'Reference Voltage' ,'Qgen_UNICORN', 'Pload_UNICORN', 'Qload_UNICORN', 'Reference voltage_OPENLF', 'Pgen_OPENLF', 'Qgen_OPENLF', 'voltage_regulator_on','Pload_OPENLF', 'Qload_OPENLF' ,'Area', 'Final bus type', 'Island number' ,  'Base Voltage' , 'Reference voltage' , 'target_v.1', 'max_q', 'min_q' , 'Pgen' , 'Qgen' , 'Pload' , 'Qload' , 'Eq. type' , 'Eq. type' , 'connected_component' , 'synchronous_component' , 'id_gen' , 'target_v' , 'p' , 'q' , 'p_load' , 'q_load' , 'U_diff' , 'theta_diff' , 'State']
                merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns], inplace = True)
                final_df, timestamp = final_columns_rename_buses(merged_df, timestamp)
                # Drop rows where all columns except 'Timestamp' and 'Bus' contain zeros
                columns_to_check = final_df.columns.difference(['Timestamp', 'Bus'])
                final_df = final_df.loc[~(final_df[columns_to_check] == 0).all(axis=1)]
                all_sheets_data['Nodes'].append(final_df)
                
                    
                #X-Nodes
                df1 = load_data(df1_path, 'Bus')
                df2 = load_data(df2_path, 'X-Nodes')    
                df1, df2 = rename_X_Nodes_data(df1, df2)
                merged_df = merge_common_data(df1, df2, merge_columns=['id'])
                merged_df = calculate_Nodes_differencies(merged_df)
                columns_to_drop = ['Bus type', 'Area', 'BUS' , 'boundary_p' , 'boundary_q' , 'v_mag' , 'v_angle' , 'I_limit'  ,'Final bus type', 'Island number' ,  'Base Voltage' , 'Reference voltage' , 'Pgen' , 'Qgen' , 'Pload' , 'Qload' , 'Eq. type' , 'Eq. type' , 'bus_id' , 'I' , 'P' , 'Q', 'boundary_p' , 'boundary_q' , 'connected_component' , 'synchronous_component' , 'id_gen' , 'target_v' , 'p' , 'q' , 'p_load' , 'q_load' , 'U_diff' , 'theta_diff' , 'State' ]     
                merged_df = merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns])
                final_df, timestamp = final_columns_rename_buses(merged_df, timestamp)
                #DROP ROWS WITH ZERO COLUMNS
                columns_to_check = final_df.columns.difference(['Timestamp', 'id'])
                final_df = final_df.loc[~(final_df[columns_to_check] == 0).all(axis=1)]
                all_sheets_data['X-Nodes'].append(final_df)                           
        else:
            
            continue  # Continue to the next loop iteration

    # Once all data is collected, save it to the Excel file in different sheets
    with pd.ExcelWriter(combined_output_path, engine='openpyxl') as writer:
        for sheet_name, dataframes in all_sheets_data.items():
            if dataframes:
                consolidated_df = pd.concat(dataframes, ignore_index=True)
                if not consolidated_df.empty:
                    consolidated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                   print(f"Sheet {sheet_name} has no data. Skipping sheet.")
            else:
                   print(f"No data for sheet {sheet_name}.")
                

    for sheet in all_sheets_data.keys():
        if sheet in ['Lines', 'X-lines']:
            # Use the function for lines
            make_adjustements_lines_to_excel(combined_output_path, sheet)
        elif sheet in ['Nodes', 'X-Nodes']:
            # Use the function for nodes
            make_adjustements_nodes_to_excel(combined_output_path, sheet)
            

if __name__ == "__main__":
    # User inforamtion
    destination_folder, destination_folder_1, Date, File_type, country_code, numbers, timestamps = get_user_inputs() 
    #Processing User's info for TCC 
    process_files_and_accumulate_data(timestamps, numbers, Date, File_type, country_code, destination_folder, destination_folder_1)

